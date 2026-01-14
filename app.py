import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import base64
import glob
import os
import json

# --- Default Configuration ---
DEFAULT_MAPPING_RULES = {
    "Service of Unit": ["I8"],
    "Item No.": ["AV8"],
    "Size": [ ["E9", "M9", "N9"] ], 
    "Type": ["Y9", "V49"], 
    "Surf/Unit (Gross/Eff)": [ ["K10", "O10", "P10"] ],
    "Fluid Name": ["T13", "AR13"], 
    "Fluid Quantity, Total": ["T14", "AR14"],
    "Temperature (In/Out)": [ 
        {"action": "vertical", "cells": ["T20", "AF20"]}, 
        {"action": "vertical", "cells": ["AR20", "BD20"]} 
    ],
    "Inlet Pressure": ["AB28", "AZ28"],
    "Velocity": ["AB29", "AZ29"],
    "Pressure Drop, Allow/Calc": [ 
        {"action": "vertical", "cells": ["T30", "AF30"]}, 
        {"action": "vertical", "cells": ["AR30", "BD30"]} 
    ],
    "Heat Exchanged": ["M32"],
    "MTD (Corrected)": ["BB32"],
    "Transfer Rate, Service": ["M33"],
    "Clean": ["AH33"],
    "Actual": ["BB33"],
    "Design/Test Pressure": ["T36"],
    "Design Temperature": ["T37"],
    "No Passes per Shell": ["T38"],
    "Tube No.": ["F43"],
    "OD": ["N43", "AC45"], 
    "Thk(Avg)": ["AC43"],
    "Length": ["AR43"],
    "Pitch": ["BG43"],
    "Tube Type": ["F44"],
    "Material": ["AH44"],
    "Tube pattern": ["BM44"],
    "Shell": ["E45"],
    "ID": ["U45"],
    "Shell Cover": ["AU45"],
    "Channel or Bonnet": ["K46"],
    "Channel Cover": ["AU46"],
    "Tubesheet-Stationary": ["K47"],
    "Tubesheet-Floating": ["AW47"],
    "Floating Head Cover": ["K48"],
    "Impingement Plate": ["AW48"],
    "Baffles-Cross": ["H49"],
    "%Cut (Diam)": ["AM49"],
    "Spacing(c/c)": ["AX49"],
    "Inlet": ["BG49"],
    "TEMA Class": ["BA57"]
}

# --- Helper Functions ---

def rules_to_df(rules_dict):
    """Convert mapping rules dict to a flat DataFrame for editing."""
    rows = []
    for label, rule_list in rules_dict.items():
        for idx, rule in enumerate(rule_list):
            row = {
                "Label": label,
                "Order": idx + 1,
                "Type": "Single",
                "Cells": ""
            }
            
            if isinstance(rule, list):
                row["Type"] = "Merge"
                row["Cells"] = ", ".join(rule)
            elif isinstance(rule, dict) and rule.get("action") == "vertical":
                row["Type"] = "Vertical"
                row["Cells"] = ", ".join(rule["cells"])
            else:
                row["Type"] = "Single"
                row["Cells"] = str(rule)
            
            rows.append(row)
    
    return pd.DataFrame(rows)

def df_to_rules(df):
    """Convert edited DataFrame back to mapping rules dict."""
    rules_dict = {}
    
    # Sort by Label and Order to ensure correct list order
    df = df.sort_values(by=["Label", "Order"])
    
    for _, row in df.iterrows():
        label = row["Label"]
        rtype = row["Type"]
        cells_str = row["Cells"]
        
        # Parse cells
        cells = [c.strip() for c in cells_str.split(",") if c.strip()]
        
        if not cells:
            continue
            
        rule = None
        if rtype == "Merge":
            rule = cells
        elif rtype == "Vertical":
            rule = {"action": "vertical", "cells": cells}
        else: # Single
            rule = cells[0]
            
        if label not in rules_dict:
            rules_dict[label] = []
        
        rules_dict[label].append(rule)
        
    return rules_dict

def find_template_file():
    if os.path.exists("template.xlsx"): return "template.xlsx"
    for file in os.listdir():
        if file.lower() == "template.xlsx": return file
    xlsx_files = [f for f in os.listdir() if f.endswith(".xlsx")]
    ignore_list = ["processed_output.xlsx", "dummy_input.xlsx", "dummy_template.xlsx"]
    candidates = [f for f in xlsx_files if f not in ignore_list and not f.startswith("dummy_") and not f.startswith("~$")]
    return candidates[0] if candidates else None

def get_cell_value(sheet, addr):
    val = sheet[addr].value
    return str(val) if val is not None else ""

def process_excel(input_file, template_file, mapping_rules):
    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet_names = input_wb.sheetnames
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        target_col_idx = 3 + i
        template_sheet.cell(row=1, column=target_col_idx).value = i + 1
        
        duplicate_counters = {key: 0 for key in mapping_rules}
        
        for row_idx in range(2, 150):
            label_cell = template_sheet.cell(row=row_idx, column=1)
            label = label_cell.value
            
            if label:
                label = str(label).strip()
                if label in mapping_rules:
                    rules = mapping_rules[label]
                    counter = duplicate_counters[label]
                    
                    if counter < len(rules):
                        rule = rules[counter]
                        try:
                            if isinstance(rule, dict) and rule.get("action") == "vertical":
                                cells = rule["cells"]
                                val1 = get_cell_value(input_sheet, cells[0])
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = val1
                                if len(cells) > 1:
                                    val2 = get_cell_value(input_sheet, cells[1])
                                    template_sheet.cell(row=row_idx + 1, column=target_col_idx).value = val2
                            elif isinstance(rule, list):
                                values = [get_cell_value(input_sheet, addr) for addr in rule]
                                merged_value = " ".join([v for v in values if v])
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = merged_value
                            else:
                                val = get_cell_value(input_sheet, rule)
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = val
                        except Exception as e:
                            print(f"Error processing {label}: {e}")
                        duplicate_counters[label] += 1

    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output

# --- Main App ---

st.set_page_config(page_title="Excel Auto-Filler", layout="wide")
st.title("Excel Data Automation App")

# Initialize Session State
if "mapping_rules" not in st.session_state:
    st.session_state.mapping_rules = DEFAULT_MAPPING_RULES

# Tabs
tab1, tab2 = st.tabs(["📂 Data Processing", "⚙️ Mapping Settings"])

# --- Tab 1: Processing ---
with tab1:
    st.markdown("### Upload & Process")
    st.info("Upload your Input File. The app will use the settings defined in the 'Mapping Settings' tab.")
    
    col1, col2 = st.columns(2)
    with col1:
        input_file = st.file_uploader("1. Upload Input File (Data)", type=['xlsx'])
    with col2:
        found_template = find_template_file()
        if found_template:
            template_file = found_template
            st.success(f"Using template: '{found_template}'")
        else:
            template_file = st.file_uploader("2. Upload Template File (Form)", type=['xlsx'])
            st.warning("No template file found. Please upload one.")

    if input_file and template_file:
        if st.button("Process Excel Files", type="primary"):
            try:
                with st.spinner("Processing..."):
                    result_file = process_excel(input_file, template_file, st.session_state.mapping_rules)
                st.success("Processing Complete!")
                st.download_button(
                    label="Download Result Excel",
                    data=result_file,
                    file_name="processed_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"An error occurred: {e}")

# --- Tab 2: Settings ---
with tab2:
    st.markdown("### ⚙️ Configure Mapping Rules")
    st.markdown("""
    - **Label**: The text in Column A of your Template file.
    - **Type**: 
        - `Single`: One cell (e.g., I8).
        - `Merge`: Multiple cells combined (e.g., E9, M9).
        - `Vertical`: Split into two rows (e.g., In/Out Temp).
    - **Cells**: Cell addresses separated by comma.
    - **Order**: If a label appears multiple times, use 1, 2, 3... to specify which rule applies to which occurrence.
    """)
    
    # Convert current rules to DF
    current_df = rules_to_df(st.session_state.mapping_rules)
    
    # Data Editor
    edited_df = st.data_editor(
        current_df,
        num_rows="dynamic",
        column_config={
            "Type": st.column_config.SelectboxColumn(
                "Mapping Type",
                options=["Single", "Merge", "Vertical"],
                required=True
            ),
            "Order": st.column_config.NumberColumn(
                "Order",
                min_value=1,
                step=1,
                required=True
            )
        },
        use_container_width=True,
        hide_index=True
    )
    
    # Save Button
    if st.button("Save Settings"):
        try:
            new_rules = df_to_rules(edited_df)
            st.session_state.mapping_rules = new_rules
            st.success("Settings saved successfully! Switch to the 'Data Processing' tab to use them.")
            # Optional: Show JSON for verification
            # st.json(new_rules)
        except Exception as e:
            st.error(f"Error saving settings: {e}")

# --- Sidebar Tools ---
with st.sidebar.expander("Developer Tools"):
    st.write("Current Directory:", os.getcwd())
    st.write("Files:", os.listdir())
    st.markdown("---")
    st.write("**Template to Base64**")
    dev_template = st.file_uploader("Upload to Convert", type=['xlsx'], key="dev_u")
    if dev_template:
        b64 = base64.b64encode(dev_template.getvalue()).decode()
        st.text_area("Base64:", b64)
