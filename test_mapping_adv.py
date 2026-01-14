import pandas as pd
import openpyxl
from io import BytesIO

# Create dummy input
wb_input = openpyxl.Workbook()
ws1 = wb_input.active
ws1.title = "Sheet1"
# Test Merge
ws1["E9"] = "500"
ws1["M9"] = "x"
ws1["N9"] = "3000"
# Test Duplicates
ws1["T13"] = "Fluid1"
ws1["AR13"] = "Fluid2"
# Test Merge + Duplicates
ws1["T20"] = "100"
ws1["AF20"] = "60"
ws1["AR20"] = "32"
ws1["BD20"] = "42"

wb_input.save("dummy_input_adv.xlsx")

# Create dummy template
wb_template = openpyxl.Workbook()
ws_temp = wb_template.active
ws_temp.title = "Template"
ws_temp["A2"] = "Size"
ws_temp["A3"] = "Fluid Name"
ws_temp["A4"] = "Fluid Name"
ws_temp["A5"] = "Temperature (In/Out)"
ws_temp["A6"] = "Temperature (In/Out)"
wb_template.save("dummy_template_adv.xlsx")

# Run logic (copied from app.py)
def process_excel_test(input_file, template_file):
    MAPPING_RULES = {
        "Size": [ ["E9", "M9", "N9"] ],
        "Fluid Name": ["T13", "AR13"],
        "Temperature (In/Out)": [ ["T20", "AF20"], ["AR20", "BD20"] ],
    }
    
    def get_cell_value(sheet, address_rule):
        if isinstance(address_rule, list):
            values = []
            for addr in address_rule:
                val = sheet[addr].value
                if val is not None:
                    values.append(str(val))
            return " ".join(values)
        else:
            return sheet[address_rule].value

    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet_names = input_wb.sheetnames
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        target_col_idx = 3 + i
        template_sheet.cell(row=1, column=target_col_idx).value = i + 1
        duplicate_counters = {key: 0 for key in MAPPING_RULES}
        
        for row_idx in range(2, 10):
            label_cell = template_sheet.cell(row=row_idx, column=1)
            label = label_cell.value
            if label:
                label = str(label).strip()
                if label in MAPPING_RULES:
                    rules = MAPPING_RULES[label]
                    counter = duplicate_counters[label]
                    if counter < len(rules):
                        rule = rules[counter]
                        value = get_cell_value(input_sheet, rule)
                        template_sheet.cell(row=row_idx, column=target_col_idx).value = value
                        duplicate_counters[label] += 1
            
    template_wb.save("dummy_output_adv.xlsx")
    print("Processed successfully.")

process_excel_test("dummy_input_adv.xlsx", "dummy_template_adv.xlsx")

# Verify output
wb_out = openpyxl.load_workbook("dummy_output_adv.xlsx")
ws_out = wb_out.active
print(f"Size: {ws_out['C2'].value}") # 500 x 3000
print(f"Fluid 1: {ws_out['C3'].value}") # Fluid1
print(f"Fluid 2: {ws_out['C4'].value}") # Fluid2
print(f"Temp 1: {ws_out['C5'].value}") # 100 60
print(f"Temp 2: {ws_out['C6'].value}") # 32 42
