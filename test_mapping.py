import pandas as pd
import openpyxl
from io import BytesIO

# Create dummy input (Complex Form)
wb_input = openpyxl.Workbook()
ws1 = wb_input.active
ws1.title = "Sheet1"
ws1["I8"] = "Service_1"
ws1["E9"] = "Size_1"
ws1["N43"] = "OD_Tube_1"
ws1["AC45"] = "OD_Shell_1"

ws2 = wb_input.create_sheet("Sheet2")
ws2["I8"] = "Service_2"
ws2["E9"] = "Size_2"
ws2["N43"] = "OD_Tube_2"
ws2["AC45"] = "OD_Shell_2"

wb_input.save("dummy_input_mapped.xlsx")

# Create dummy template (List)
wb_template = openpyxl.Workbook()
ws_temp = wb_template.active
ws_temp.title = "Template"
ws_temp["A2"] = "Service of Unit"
ws_temp["A3"] = "Size"
ws_temp["A4"] = "OD" # Tube
ws_temp["A5"] = "OD" # Shell
wb_template.save("dummy_template_mapped.xlsx")

# Run logic (copied from app.py for testing)
def process_excel_test(input_file, template_file):
    MAPPING = {
        "Service of Unit": "I8",
        "Size": "E9",
    }
    MAPPING_LIST = {
        "OD": ["N43", "AC45"]
    }
    
    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet_names = input_wb.sheetnames
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        target_col_idx = 3 + i
        template_sheet.cell(row=1, column=target_col_idx).value = i + 1
        duplicate_counters = {key: 0 for key in MAPPING_LIST}
        
        for row_idx in range(2, 10):
            label_cell = template_sheet.cell(row=row_idx, column=1)
            label = label_cell.value
            if label:
                label = str(label).strip()
                cell_address = None
                if label in MAPPING_LIST:
                    counter = duplicate_counters[label]
                    if counter < len(MAPPING_LIST[label]):
                        cell_address = MAPPING_LIST[label][counter]
                        duplicate_counters[label] += 1
                elif label in MAPPING:
                    cell_address = MAPPING[label]
                
                if cell_address:
                    value = input_sheet[cell_address].value
                    template_sheet.cell(row=row_idx, column=target_col_idx).value = value
            
    template_wb.save("dummy_output_mapped.xlsx")
    print("Processed successfully.")

process_excel_test("dummy_input_mapped.xlsx", "dummy_template_mapped.xlsx")

# Verify output
wb_out = openpyxl.load_workbook("dummy_output_mapped.xlsx")
ws_out = wb_out.active
print(f"Col C (Sheet1) Service: {ws_out['C2'].value}") # Service_1
print(f"Col C (Sheet1) OD Tube: {ws_out['C4'].value}") # OD_Tube_1
print(f"Col C (Sheet1) OD Shell: {ws_out['C5'].value}") # OD_Shell_1
print(f"Col D (Sheet2) Service: {ws_out['D2'].value}") # Service_2
