import pandas as pd
import openpyxl
from io import BytesIO

# Create dummy input
wb_input = openpyxl.Workbook()
ws1 = wb_input.active
ws1.title = "Sheet1"
ws1["T20"] = "100"
ws1["AF20"] = "60"
ws1["AR20"] = "32"
ws1["BD20"] = "42"
wb_input.save("dummy_input_vert.xlsx")

# Create dummy template
wb_template = openpyxl.Workbook()
ws_temp = wb_template.active
ws_temp.title = "Template"
ws_temp["A5"] = "Temperature (In/Out)"
# A6 is empty
ws_temp["A7"] = "Temperature (In/Out)"
# A8 is empty
wb_template.save("dummy_template_vert.xlsx")

# Run logic (copied from app.py)
def process_excel_test(input_file, template_file):
    MAPPING_RULES = {
        "Temperature (In/Out)": [ 
            {"action": "vertical", "cells": ["T20", "AF20"]}, 
            {"action": "vertical", "cells": ["AR20", "BD20"]} 
        ],
    }
    
    def get_cell_value(sheet, addr):
        val = sheet[addr].value
        return str(val) if val is not None else ""

    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet_names = input_wb.sheetnames
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        target_col_idx = 3 + i
        template_sheet.cell(row=1, column=target_col_idx).value = i + 1
        duplicate_counters = {key: 0 for key in MAPPING_RULES}
        
        for row_idx in range(2, 10):
            label_cell = template_sheet.cell(row=row_idx, column=1)
            label = label_cell.value
            if label:
                label = str(label).strip()
                if label in MAPPING_RULES:
                    rules = MAPPING_RULES[label]
                    counter = duplicate_counters[label]
                    if counter < len(rules):
                        rule = rules[counter]
                        if isinstance(rule, dict) and rule.get("action") == "vertical":
                            cells = rule["cells"]
                            val1 = get_cell_value(input_sheet, cells[0])
                            template_sheet.cell(row=row_idx, column=target_col_idx).value = val1
                            if len(cells) > 1:
                                val2 = get_cell_value(input_sheet, cells[1])
                                template_sheet.cell(row=row_idx + 1, column=target_col_idx).value = val2
                        duplicate_counters[label] += 1
            
    template_wb.save("dummy_output_vert.xlsx")
    print("Processed successfully.")

process_excel_test("dummy_input_vert.xlsx", "dummy_template_vert.xlsx")

# Verify output
wb_out = openpyxl.load_workbook("dummy_output_vert.xlsx")
ws_out = wb_out.active
print(f"Temp 1 In (Row 5): {ws_out['C5'].value}") # 100
print(f"Temp 1 Out (Row 6): {ws_out['C6'].value}") # 60
print(f"Temp 2 In (Row 7): {ws_out['C7'].value}") # 32
print(f"Temp 2 Out (Row 8): {ws_out['C8'].value}") # 42
