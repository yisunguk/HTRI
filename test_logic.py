import pandas as pd
import openpyxl
from io import BytesIO

# Create dummy input
wb_input = openpyxl.Workbook()
ws1 = wb_input.active
ws1.title = "Sheet1"
# Fill C2:C48
for i in range(2, 49):
    ws1.cell(row=i, column=3).value = f"Data1_{i}"

ws2 = wb_input.create_sheet("Sheet2")
for i in range(2, 49):
    ws2.cell(row=i, column=3).value = f"Data2_{i}"

wb_input.save("dummy_input.xlsx")

# Create dummy template
wb_template = openpyxl.Workbook()
ws_temp = wb_template.active
ws_temp.title = "TemplateMaster"
ws_temp.cell(row=1, column=1).value = "Header"
wb_template.save("dummy_template.xlsx")

# Run logic (copied from app.py for testing)
def process_excel_test(input_file, template_file):
    template_wb = openpyxl.load_workbook(template_file)
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    
    output_sheets = []
    input_sheet_names = input_wb.sheetnames
    base_template_sheet = template_wb.active
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        if i == 0:
            target_sheet = base_template_sheet
            target_sheet.title = sheet_name
        else:
            target_sheet = template_wb.copy_worksheet(base_template_sheet)
            target_sheet.title = sheet_name
        
        for row_offset in range(47):
            input_row = 2 + row_offset
            output_row = 3 + row_offset
            value = input_sheet.cell(row=input_row, column=3).value
            target_sheet.cell(row=output_row, column=3).value = value
            
    template_wb.save("dummy_output.xlsx")
    print("Processed successfully.")

process_excel_test("dummy_input.xlsx", "dummy_template.xlsx")

# Verify output
wb_out = openpyxl.load_workbook("dummy_output.xlsx")
print(f"Output sheets: {wb_out.sheetnames}")
ws_out1 = wb_out["Sheet1"]
print(f"Sheet1 C3: {ws_out1['C3'].value}") # Should be Data1_2
print(f"Sheet1 C49: {ws_out1['C49'].value}") # Should be Data1_48
ws_out2 = wb_out["Sheet2"]
print(f"Sheet2 C3: {ws_out2['C3'].value}") # Should be Data2_2
